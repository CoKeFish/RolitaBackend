import pandas as pd
import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def obtener_datos_por_semana(db_path):
    """
    Consulta la base de datos SQLite para obtener los datos de los registros de los buses organizados por semana.
    """
    # Conectar a la base de datos
    conn = sqlite3.connect(db_path)

    # Query para obtener datos de sensores
    query_sensores = """
    SELECT time, bus
    FROM sensores
    ORDER BY time
    """

    # Query para obtener datos de P60
    query_P60 = """
    SELECT fechaHoraLecturaDato AS time, substr(idVehiculo, -4, 4) AS bus
    FROM P60
    ORDER BY fechaHoraLecturaDato
    """

    # Leer los datos de ambas tablas
    df_sensores = pd.read_sql_query(query_sensores, conn)
    df_P60 = pd.read_sql_query(query_P60, conn)

    conn.close()

    # Convertir las columnas de fecha a tipo datetime
    df_sensores['time'] = pd.to_datetime(df_sensores['time'], format='mixed')
    df_P60['time'] = pd.to_datetime(df_P60['time'], format='mixed')

    # Extraer los últimos 4 dígitos del bus para alinearlo con idVehiculo
    df_sensores['bus'] = df_sensores['bus'].astype(str).str[-4:]

    # Agregar columna de semana, año y día de la semana a ambos DataFrames
    for df in [df_sensores, df_P60]:
        df['Semana'] = df['time'].apply(lambda x: x.strftime('%U'))
        df['Año'] = df['time'].apply(lambda x: x.strftime('%Y'))
        df['DíaSemana'] = df['time'].apply(lambda x: x.strftime('%A'))

    # Agregar etiquetas para identificar la fuente de los datos
    df_sensores['source'] = 'TXT'
    df_P60['source'] = 'LOG'

    # Concatenar ambos DataFrames
    combined_df = pd.concat([df_sensores, df_P60])

    # Agrupar por Año, Semana, Día de la semana, y bus
    grouped = combined_df.groupby(['Año', 'Semana', 'DíaSemana', 'bus', 'source']).agg(
        HoraInicio=('time', 'min'),
        HoraFin=('time', 'max')
    ).reset_index()

    # Crear columna combinada para mostrar la información en cada celda sin repetir el bus
    def format_info(row):
        if row['source'] == 'TXT':
            return f"({row['HoraInicio'].strftime('%H:%M')} - {row['HoraFin'].strftime('%H:%M')}) [TXT]"
        else:  # source is 'LOG'
            return "[LOG]"

    grouped['info'] = grouped.apply(format_info, axis=1)

    return grouped


def crear_informe_excel(db_path, output_file):
    """
    Crea un informe en Excel a partir de los datos de la base de datos SQLite.
    """
    datos = obtener_datos_por_semana(db_path)

    # Crear un archivo de Excel con hojas por cada año
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for año in datos['Año'].unique():
            datos_año = datos[datos['Año'] == año]

            # Crear una tabla pivotada donde las filas son combinaciones de semana y bus
            pivot_table = pd.pivot_table(
                datos_año,
                index=['Semana', 'bus'],  # Índice combinado de semana y bus
                columns='DíaSemana',
                values='info',  # Usar la columna combinada 'info'
                aggfunc=lambda x: '\n'.join(x),  # Combinar múltiples entradas por día
                fill_value=''  # Rellenar valores vacíos con cadenas vacías
            )

            # Ordenar columnas de días de la semana
            dias_ordenados = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            pivot_table = pivot_table.reindex(columns=dias_ordenados)

            # Eliminar semanas sin datos (no se dejarán en blanco)
            pivot_table.dropna(how='all', inplace=True)

            # Guardar cada tabla en una hoja de Excel
            pivot_table.to_excel(writer, sheet_name=str(año))

    # Cargar el archivo Excel guardado para aplicar estilos de color
    workbook = load_workbook(output_file)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # Definir colores de relleno
        color_naranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Naranja
        color_amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Amarillo
        color_verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Verde

        # Aplicar color a las celdas basadas en la presencia de datos en 'sensores' y 'P60'
        for row in sheet.iter_rows(min_row=2, min_col=2):
            for cell in row:
                if cell.value:  # Solo colorear celdas no vacías
                    contains_txt = '[TXT]' in cell.value
                    contains_log = '[LOG]' in cell.value
                    if contains_txt and contains_log:
                        cell.fill = color_verde
                    elif contains_txt:
                        cell.fill = color_naranja
                    elif contains_log:
                        cell.fill = color_amarillo

    # Guardar los cambios en el archivo
    workbook.save(output_file)

    print(f"Informe guardado en {output_file} con celdas coloreadas.")


if __name__ == "__main__":
    db_path = "C:\\Users\\rrtc2\\Documents\\mi_base_de_datos.db"  # Ruta de tu base de datos
    output_file = "C:\\Users\\rrtc2\\Documents\\informe_semanal.xlsx"  # Ruta donde se guardará el informe

    # Crear el informe en Excel
    crear_informe_excel(db_path, output_file)
